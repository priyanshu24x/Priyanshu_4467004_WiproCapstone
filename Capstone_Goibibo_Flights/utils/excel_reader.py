# Reads Excel files (.xlsx) and returns data as a list of dictionaries (each row becomes a dictionary with column headers as keys).

import os
import sys
import openpyxl

class ExcelReader:

    @staticmethod
    def read_excel(file_name, sheet_name):
        excel_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            "data",
            file_name
        )

        data = []
        try:
            # Load the workbook, using data_only=True to get calculated values instead of formulas
            # Pass the full excel_path, not just the file_name
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = workbook[sheet_name]
        except FileNotFoundError:
            print(f"Error: Excel file not found at '{excel_path}'", file=sys.stderr)
            return []
        except KeyError:
            print(f"Error: Sheet '{sheet_name}' not found in '{file_name}'", file=sys.stderr)
            return []
        except Exception as e:
            print(f"An unexpected error occurred while loading the Excel file: {e}", file=sys.stderr)
            return []

        # Extract headers from the first row
        headers = [cell.value for cell in sheet[1]]
        # Filter out None values from headers if necessary, or handle them
        # For example, to replace None with an empty string:
        # headers = [cell.value if cell.value is not None else "" for cell in sheet[1]]

        # Iterate over rows starting from the second row (skipping headers)
        for row_index in range(2, sheet.max_row + 1):
            row_data = {}
            for col_index in range(1, sheet.max_column + 1):
                header = headers[col_index - 1]
                row_data[header] = sheet.cell(row=row_index, column=col_index).value
            data.append(row_data)

        return data