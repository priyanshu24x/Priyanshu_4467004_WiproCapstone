# Reads Excel files (.xlsx) and returns data as a list of dictionaries (each row becomes a dictionary with column headers as keys).
import os
import sys
import openpyxl


class ExcelReader:

    @staticmethod
    def read_excel(file_name, sheet_name):
        excel_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            "data",
            file_name
        )

        data = []
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = workbook[sheet_name]
        except FileNotFoundError:
            print(f"Error: Excel file not found at '{excel_path}'", file=sys.stderr)
            return []
        except KeyError:
            print(f"Error: Sheet '{sheet_name}' not found in '{file_name}'", file=sys.stderr)
            return []
        except Exception as e:
            print(f"An unexpected error occurred while loading the Excel file: {e}", file=sys.stderr)
            return []

        headers = [cell.value for cell in sheet[1]]

        for row_index in range(2, sheet.max_row + 1):
            row_data = {}
            for col_index in range(1, sheet.max_column + 1):
                header = headers[col_index - 1]
                raw_value = sheet.cell(row=row_index, column=col_index).value

                # ==========================================
                # SANITIZATION MECHANISM
                # ==========================================
                if raw_value is not None:
                    clean_value = str(raw_value).strip()
                    # Strip Excel's floating point zeros (e.g., "26.0" -> "26")
                    if clean_value.endswith('.0'):
                        clean_value = clean_value[:-2]
                    # Format the month perfectly for the XPath
                    if header == 'TravelMonth':
                        clean_value = clean_value.capitalize()

                    row_data[header] = clean_value
                else:
                    row_data[header] = ""

            data.append(row_data)

        return data